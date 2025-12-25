"""
Excel dosyası işlemleri için yardımcı modül
Form verilerini Excel'den okur ve günceller
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Excel dosyası yolu - script'in çalıştığı dizinde
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "form_data.xlsx")

def create_default_excel():
    """Default değerlerle Excel dosyası oluşturur"""
    wb = Workbook()
    
    # Varsayılan sheet'i sil
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Vehicles sheet
    ws_vehicles = wb.create_sheet("Vehicles")
    ws_vehicles.append(["Vehicle"])
    ws_vehicles.append(["SPRINTER: BT-48331"])
    ws_vehicles.append(["RAM-Promaster 2500 (2021)"])
    ws_vehicles.append(["MERCEDES-2500 Cargo Van (2013)"])
    
    # FuelLevels sheet
    ws_fuel = wb.create_sheet("FuelLevels")
    ws_fuel.append(["Level"])
    ws_fuel.append(["Full"])
    ws_fuel.append(["3/4"])
    ws_fuel.append(["Half"])
    ws_fuel.append(["1/4"])
    ws_fuel.append(["Other"])
    
    # ExteriorChecks sheet
    ws_exterior = wb.create_sheet("ExteriorChecks")
    ws_exterior.append(["Field"])
    exterior_fields = ["headlights", "break_lights", "indicators", "mirrors", "windows", 
                      "windshield", "wiper_fluid", "wipers", "tires", "body_paint"]
    for field in exterior_fields:
        ws_exterior.append([field])
    
    # EngineChecks sheet
    ws_engine = wb.create_sheet("EngineChecks")
    ws_engine.append(["Field"])
    engine_fields = ["engine_noise", "coolant_level", "brake_fluid", "battery_condition", "belts_hoses"]
    for field in engine_fields:
        ws_engine.append([field])
    
    # SafetyEquipment sheet
    ws_safety = wb.create_sheet("SafetyEquipment")
    ws_safety.append(["Field"])
    safety_fields = ["seatbelts", "fire_extinguisher", "first_aid_kit", "horn", "jumper_cables", "snow_brush"]
    for field in safety_fields:
        ws_safety.append([field])
    
    # InteriorChecks sheet
    ws_interior = wb.create_sheet("InteriorChecks")
    ws_interior.append(["Field"])
    interior_fields = ["cleanliness", "dashboard_lights", "hvac", "seats"]
    for field in interior_fields:
        ws_interior.append([field])
    
    # Items sheet
    ws_items = wb.create_sheet("Items")
    ws_items.append(["Item"])
    ws_items.append(["Fuel Card"])
    ws_items.append(["Measuring Tape"])
    ws_items.append(["Safety Vest"])
    
    # Users sheet
    ws_users = wb.create_sheet("Users")
    ws_users.append(["Username", "Password", "Full Name"])
    ws_users.append(["innovodriver", "123456", "Mehmet Berk"])
    
    wb.save(EXCEL_FILE)
    return wb

def get_excel_file():
    """Excel dosyasını açar, yoksa oluşturur"""
    if not os.path.exists(EXCEL_FILE):
        return create_default_excel()
    return load_workbook(EXCEL_FILE)

def load_vehicles():
    """Vehicles sheet'inden araç listesini okur"""
    wb = get_excel_file()
    ws = wb["Vehicles"]
    vehicles = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            vehicles.append(row[0])
    return vehicles

def load_fuel_levels():
    """FuelLevels sheet'inden yakıt seviyelerini okur"""
    wb = get_excel_file()
    ws = wb["FuelLevels"]
    levels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            levels.append(row[0])
    return levels

def load_check_fields(category):
    """İlgili sheet'ten kontrolleri okur
    category: 'ExteriorChecks', 'EngineChecks', 'SafetyEquipment', 'InteriorChecks'
    """
    wb = get_excel_file()
    if category not in wb.sheetnames:
        return []
    ws = wb[category]
    fields = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            fields.append(row[0])
    return fields

def load_items():
    """Items sheet'inden eşya listesini okur"""
    wb = get_excel_file()
    ws = wb["Items"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            items.append(row[0])
    return items

def load_users():
    """Users sheet'inden kullanıcıları okur"""
    wb = get_excel_file()
    ws = wb["Users"]
    users = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[2]:
            users[row[0]] = {
                "password": row[1],
                "full_name": row[2]
            }
    return users

def add_vehicle(vehicle_name):
    """Yeni araç ekler"""
    wb = get_excel_file()
    ws = wb["Vehicles"]
    ws.append([vehicle_name])
    wb.save(EXCEL_FILE)

def add_fuel_level(level):
    """Yeni yakıt seviyesi ekler"""
    wb = get_excel_file()
    ws = wb["FuelLevels"]
    ws.append([level])
    wb.save(EXCEL_FILE)

def add_check_field(category, field_name):
    """Yeni kontrol alanı ekler"""
    wb = get_excel_file()
    if category not in wb.sheetnames:
        wb.create_sheet(category)
        ws = wb[category]
        ws.append(["Field"])
    else:
        ws = wb[category]
    ws.append([field_name])
    wb.save(EXCEL_FILE)

def add_item(item_name):
    """Yeni eşya ekler"""
    wb = get_excel_file()
    ws = wb["Items"]
    ws.append([item_name])
    wb.save(EXCEL_FILE)

def add_user(username, password, full_name):
    """Yeni kullanıcı ekler"""
    wb = get_excel_file()
    ws = wb["Users"]
    ws.append([username, password, full_name])
    wb.save(EXCEL_FILE)

# Form gönderimleri için ayrı Excel dosyası
SUBMISSIONS_FILE = os.path.join(os.path.dirname(__file__), "form_submissions.xlsx")

def save_form_submission(form_data):
    """Form verilerini ayrı bir Excel dosyasına kaydeder"""
    from datetime import datetime
    from openpyxl import Workbook
    
    # Dosya varsa aç, yoksa yeni oluştur
    if os.path.exists(SUBMISSIONS_FILE):
        wb = load_workbook(SUBMISSIONS_FILE)
        if "Submissions" not in wb.sheetnames:
            ws = wb.create_sheet("Submissions")
        else:
            ws = wb["Submissions"]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        ws = wb.create_sheet("Submissions")
        # Başlık satırı oluştur
        headers = [
            "Timestamp", "Driver Name", "Vehicle", "Odometer Start", 
            "Fuel Level", "Oil Level", "Fuel Card", "Measuring Tape", 
            "Safety Vest", "Fuel Amount"
        ]
        
        # Exterior checks başlıkları
        exterior_fields = load_check_fields("ExteriorChecks")
        for field in exterior_fields:
            headers.append(f"Exterior_{field}")
        
        # Engine checks başlıkları
        engine_fields = load_check_fields("EngineChecks")
        for field in engine_fields:
            headers.append(f"Engine_{field}")
        
        # Safety equipment başlıkları
        safety_fields = load_check_fields("SafetyEquipment")
        for field in safety_fields:
            headers.append(f"Safety_{field}")
        
        # Interior checks başlıkları
        interior_fields = load_check_fields("InteriorChecks")
        for field in interior_fields:
            headers.append(f"Interior_{field}")
        
        ws.append(headers)
    
    # Veri satırı oluştur
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        form_data.get("driver_name", ""),
        form_data.get("vehicle", ""),
        form_data.get("odometer_start", ""),
        form_data.get("fuel_level", ""),
        form_data.get("oil_level", ""),
        form_data.get("fuel_card", ""),
        form_data.get("measuring_tape", ""),
        form_data.get("safety_vest", ""),
        form_data.get("fuel_amount", "")
    ]
    
    # Exterior checks değerleri
    exterior_checks = form_data.get("exterior_checks", {})
    for field in load_check_fields("ExteriorChecks"):
        row.append(exterior_checks.get(field, ""))
    
    # Engine checks değerleri
    engine_checks = form_data.get("engine_checks", {})
    for field in load_check_fields("EngineChecks"):
        row.append(engine_checks.get(field, ""))
    
    # Safety equipment değerleri
    safety_checks = form_data.get("safety_checks", {})
    for field in load_check_fields("SafetyEquipment"):
        row.append(safety_checks.get(field, ""))
    
    # Interior checks değerleri
    interior_checks = form_data.get("interior_checks", {})
    for field in load_check_fields("InteriorChecks"):
        row.append(interior_checks.get(field, ""))
    
    ws.append(row)
    wb.save(SUBMISSIONS_FILE)

